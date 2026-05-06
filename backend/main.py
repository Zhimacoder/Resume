"""
智能简历筛选工具 - FastAPI后端主入口
"""

import os
import io
import json
import asyncio
from pathlib import Path
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
import uvicorn

# 导入业务模块
from config import load_config, save_config, get_config_status, DEFAULT_CONFIG
from parser import parse_document
from ocr import is_image_file, process_image
from desensitizer import desensitize_resume
from llm import analyze_resume, test_connection


# 创建FastAPI应用
app = FastAPI(title="智能简历筛选工具", version="1.2.0")

# 配置CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 获取项目根目录
ROOT_DIR = Path(__file__).parent.parent
FRONTEND_DIR = ROOT_DIR / "frontend"

# 挂载静态文件
if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


# ==================== 页面路由 ====================

@app.get("/", response_class=HTMLResponse)
async def index_page():
    """主页"""
    index_file = FRONTEND_DIR / "index.html"
    if index_file.exists():
        return FileResponse(str(index_file))
    return HTMLResponse(content="<h1>智能简历筛选工具</h1><p>index.html not found</p>")


@app.get("/result", response_class=HTMLResponse)
async def result_page():
    """结果页"""
    result_file = FRONTEND_DIR / "result.html"
    if result_file.exists():
        return FileResponse(str(result_file))
    return HTMLResponse(content="<h1>结果页</h1><p>result.html not found</p>")


@app.get("/config", response_class=HTMLResponse)
async def config_page():
    """API配置页"""
    config_file = FRONTEND_DIR / "config.html"
    if config_file.exists():
        return FileResponse(str(config_file))
    return HTMLResponse(content="<h1>API配置页</h1><p>config.html not found</p>")


# ==================== API配置接口 ====================

@app.get("/api/config/status")
async def api_get_config_status():
    """获取配置状态"""
    return get_config_status()


@app.get("/api/config")
async def api_get_config():
    """获取完整配置"""
    return load_config()


@app.post("/api/config")
async def api_save_config(config_data: Dict[str, Any]):
    """保存配置"""
    success = save_config(config_data)
    if success:
        return {"success": True, "message": "配置保存成功"}
    else:
        raise HTTPException(status_code=500, detail="配置保存失败")


# ==================== API连通性测试 ====================

@app.post("/api/test-connection")
async def api_test_connection(payload: Dict[str, Any]):
    """
    测试API Key连通性

    请求体：
    {
        "model_type": "doubao",
        "api_key": "xxx",
        "endpoint": "https://...",
        "model_name": "doubao-pro-32k"  // 可选
    }
    """
    model_type = payload.get("model_type", "")
    api_key = payload.get("api_key", "")
    endpoint = payload.get("endpoint", "")
    model_name = payload.get("model_name", "")

    if not api_key:
        raise HTTPException(status_code=400, detail="api_key 不能为空")
    if not endpoint:
        raise HTTPException(status_code=400, detail="endpoint 不能为空")

    result = test_connection(model_type, api_key, endpoint, model_name)
    return result


# ==================== 简历筛选接口 ====================

@app.post("/api/screening")
async def api_screening(
    jd_content: str = Form(...),
    files: List[UploadFile] = File(...),
    dimensions: Optional[str] = Form(None)
):
    """
    智能筛选主流程接口

    流程：
    1. 解析上传的简历文件
    2. 对图片简历进行OCR识别
    3. 对简历内容进行脱敏
    4. 调用大模型分析匹配度
    5. 返回结构化结果

    dimensions（可选JSON字符串）：
    {
        "work_years": "3年以上",
        "education": "本科及以上",
        "skill_weights": "Python、机器学习优先",
        "extra": "有大厂经验优先"
    }
    """
    # 获取配置
    config = load_config()

    # 检查配置有效性
    if not config.get("is_config_valid", False):
        raise HTTPException(status_code=400, detail="请先完成APIkey配置")

    current_model = config.get("current_model", "doubao")
    model_config = config.get("models", {}).get(current_model, {})

    api_key = model_config.get("api_key", "")
    endpoint = model_config.get("endpoint", "")
    model_name = model_config.get("model_version", "")

    if not api_key or not endpoint:
        raise HTTPException(status_code=400, detail="当前模型配置不完整")

    # 解析自定义匹配维度
    dimensions_dict = None
    if dimensions:
        try:
            dimensions_dict = json.loads(dimensions)
        except json.JSONDecodeError:
            dimensions_dict = None

    results = []
    errors = []

    # 处理每个简历文件
    for file in files:
        try:
            # 读取文件内容
            file_content = await file.read()
            filename = file.filename

            # 解析简历
            resume_text = ""

            # 根据文件类型选择解析方式
            ext = Path(filename).suffix.lower()

            if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp']:
                # 图片文件 - 使用OCR
                resume_text = process_image(file_content, filename)
            else:
                # 文档文件 - 解析内容
                resume_text = parse_document(file_content, filename)

            if not resume_text or "[解析失败" in resume_text or "[需要" in resume_text:
                errors.append({
                    "filename": filename,
                    "error": f"简历解析失败: {resume_text}"
                })
                continue

            # 脱敏处理
            desensitized_text = desensitize_resume(resume_text)

            # 调用大模型分析
            analysis_result = analyze_resume(
                model_type=current_model,
                api_key=api_key,
                endpoint=endpoint,
                jd_content=jd_content,
                resume_content=desensitized_text,
                resume_name=filename,
                model_name=model_name,
                dimensions=dimensions_dict
            )

            if analysis_result:
                results.append(analysis_result)
            else:
                errors.append({
                    "filename": filename,
                    "error": "大模型分析失败"
                })

        except Exception as e:
            errors.append({
                "filename": file.filename if file else "unknown",
                "error": str(e)
            })

    # 按匹配度排序
    results.sort(key=lambda x: x.get("score", 0), reverse=True)

    return {
        "success": True,
        "total": len(files),
        "results": results,
        "errors": errors,
        "model_used": current_model,
        "model_name": model_config.get("model_name", current_model)
    }


@app.post("/api/parse-resume")
async def api_parse_resume(file: UploadFile = File(...)):
    """
    单独解析简历接口（用于预览）
    """
    try:
        file_content = await file.read()
        filename = file.filename

        # 解析简历
        ext = Path(filename).suffix.lower()

        if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp']:
            resume_text = process_image(file_content, filename)
        else:
            resume_text = parse_document(file_content, filename)

        # 脱敏
        desensitized_text = desensitize_resume(resume_text)

        return {
            "success": True,
            "original_text": resume_text,
            "desensitized_text": desensitized_text,
            "filename": filename
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"简历解析失败: {str(e)}")


# ==================== Excel导出接口 ====================

@app.post("/api/export")
async def api_export_excel(payload: Dict[str, Any]):
    """
    导出筛选结果为Excel文件

    请求体：
    {
        "results": [...],   // 筛选结果数组
        "jd_summary": "xx"  // 可选，JD摘要用作Sheet标题备注
    }
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="导出功能需要安装openpyxl库，请运行: pip install openpyxl")

    results = payload.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="没有可导出的数据")

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "简历筛选结果"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_align = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill(start_color="F0F0FF", end_color="F0F0FF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )

    # 表头
    headers = [
        ("排名", 6),
        ("简历文件名", 25),
        ("匹配度评分", 12),
        ("简历概述", 40),
        ("匹配点", 45),
        ("不足点", 35),
        ("建议面试问题", 50),
    ]

    for col_idx, (header_text, col_width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 28

    # 填充数据行
    for row_idx, item in enumerate(results, start=2):
        rank = row_idx - 1
        resume_name = item.get("resume_name", "")
        score = item.get("score", 0)
        summary = item.get("summary", "")
        matching_points = "\n".join(
            f"• {p}" for p in (item.get("matching_points") or [])
        )
        shortcomings = "\n".join(
            f"• {s}" for s in (item.get("shortcomings") or [])
        )
        interview_qs = "\n".join(
            f"{i+1}. {q}" for i, q in enumerate(item.get("interview_questions") or [])
        )

        row_data = [rank, resume_name, score, summary, matching_points, shortcomings, interview_qs]
        use_alt = (row_idx % 2 == 0)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            if use_alt:
                cell.fill = alt_fill
            if col_idx == 3:
                # 评分列居中加粗
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.font = Font(bold=True, size=12)

        # 自动行高（近似估算）
        max_lines = max(
            len(str(v).split("\n")) for v in row_data if v
        ) if any(row_data) else 1
        ws.row_dimensions[row_idx].height = max(18, min(max_lines * 16, 200))

    # 冻结首行
    ws.freeze_panes = "A2"

    # 输出到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import datetime
    filename = f"简历筛选结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


# ==================== 健康检查 ====================

@app.get("/api/health")
async def health_check():
    """健康检查"""
    return {
        "status": "ok",
        "version": "1.2.0"
    }


# ==================== 启动配置 ====================

if __name__ == "__main__":
    # 确保配置目录存在
    config_dir = ROOT_DIR / "config"
    config_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 50)
    print("智能简历筛选工具 V1.2")
    print("=" * 50)
    print(f"前端目录: {FRONTEND_DIR}")
    print(f"配置目录: {config_dir}")
    print("=" * 50)
    print("服务启动成功！")
    print("访问地址: http://localhost:8000")
    print("=" * 50)

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
