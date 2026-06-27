"""
智能简历筛选工具 - FastAPI后端主入口
"""

import os
import io
import json
import asyncio
from pathlib import Path
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
import uvicorn

# 导入业务模块
from config import load_config, save_config, get_config_status, get_masked_config, DEFAULT_CONFIG
from parser import parse_document
from ocr import is_image_file, process_image
from desensitizer import desensitize_resume
from age_extractor import extract_age
from llm import analyze_resume, test_connection

# 并发控制：同时处理最多 3 份简历
_MAX_CONCURRENT_RESUMES = 3

# CORS 允许来源（支持环境变量配置，逗号分隔）
_CORS_ORIGINS = os.environ.get(
    "CORS_ORIGINS",
    "http://localhost:8000,http://127.0.0.1:8000,https://zhimacoder.com"
).split(",")

# 生产模式开关：true 时强制每个用户填自己的 API Key（存浏览器 localStorage），
# 后端不读服务器端配置；false 时（默认，本地开发）沿用 load_config() 行为
REQUIRE_USER_API_KEY = os.environ.get("REQUIRE_USER_API_KEY", "false").lower() == "true"


# 创建FastAPI应用
app = FastAPI(title="智能简历筛选工具", version="1.2.0")

# 配置CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=_CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 获取项目根目录
ROOT_DIR = Path(__file__).parent.parent
FRONTEND_DIR = ROOT_DIR / "frontend"
WEBSITE_DIR = ROOT_DIR / "website"

# 挂载静态文件（业务前端页面 + website favicon 都通过此挂载服务）
if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


# ==================== 页面路由 ====================

@app.get("/", response_class=HTMLResponse)
async def landing_page():
    """官网首页（产品介绍页）"""
    landing_file = WEBSITE_DIR / "index.html"
    if landing_file.exists():
        return FileResponse(str(landing_file))
    return HTMLResponse(content="<h1>智能简历筛选工具</h1><p>website/index.html not found</p>")


@app.get("/app", response_class=HTMLResponse)
async def index_page():
    """应用主页（简历筛选工具）"""
    index_file = FRONTEND_DIR / "index.html"
    if index_file.exists():
        return FileResponse(str(index_file))
    return HTMLResponse(content="<h1>智能简历筛选工具</h1><p>index.html not found</p>")


@app.get("/result", response_class=HTMLResponse)
async def result_page():
    """结果页"""
    result_file = FRONTEND_DIR / "result.html"
    if result_file.exists():
        return FileResponse(str(result_file))
    return HTMLResponse(content="<h1>结果页</h1><p>result.html not found</p>")


@app.get("/config", response_class=HTMLResponse)
async def config_page():
    """API配置页"""
    config_file = FRONTEND_DIR / "config.html"
    if config_file.exists():
        return FileResponse(str(config_file))
    return HTMLResponse(content="<h1>API配置页</h1><p>config.html not found</p>")


# ==================== API配置接口 ====================

@app.get("/api/config/status")
async def api_get_config_status():
    """获取配置状态"""
    if REQUIRE_USER_API_KEY:
        return {
            "is_config_valid": False,
            "current_model": "deepseek",
            "models": {},
            "require_user_api_key": True
        }
    status = get_config_status()
    status["require_user_api_key"] = False
    return status


@app.get("/api/config")
async def api_get_config():
    """获取配置（API Key 已脱敏）"""
    if REQUIRE_USER_API_KEY:
        raise HTTPException(
            status_code=403,
            detail="生产模式已启用，API Key 由用户在浏览器端填写，服务器端不再提供配置"
        )
    return get_masked_config()


@app.post("/api/config")
async def api_save_config(config_data: Dict[str, Any]):
    """保存配置"""
    if REQUIRE_USER_API_KEY:
        raise HTTPException(
            status_code=403,
            detail="生产模式已启用，API Key 由用户在浏览器端填写，服务器端不再持久化配置"
        )
    success = save_config(config_data)
    if success:
        return {"success": True, "message": "配置保存成功"}
    else:
        raise HTTPException(status_code=500, detail="配置保存失败")


# ==================== API连通性测试 ====================

@app.post("/api/test-connection")
async def api_test_connection(payload: Dict[str, Any]):
    """
    测试API Key连通性

    请求体：
    {
        "model_type": "deepseek",
        "api_key": "xxx",
        "endpoint": "https://api.deepseek.com/chat/completions",
        "model_name": "deepseek-chat"  // 自定义模型时必填
    }
    """
    model_type = payload.get("model_type", "")
    api_key = payload.get("api_key", "")
    endpoint = payload.get("endpoint", "")
    model_name = payload.get("model_name", "")
    api_secret = payload.get("api_secret", "")

    if not api_key:
        raise HTTPException(status_code=400, detail="api_key 不能为空")
    if not endpoint:
        raise HTTPException(status_code=400, detail="endpoint 不能为空")

    result = test_connection(model_type, api_key, endpoint, model_name, api_secret)
    return result


# ==================== 简历筛选接口 ====================

async def _process_single_resume(
    file: UploadFile,
    current_model: str,
    api_key: str,
    endpoint: str,
    jd_content: str,
    model_name: str,
    dimensions_dict: Optional[Dict[str, Any]],
    api_secret: str = ""
) -> Dict[str, Any]:
    """处理单份简历：解析 → 提取年龄 → 脱敏 → LLM 分析 → 后处理（在线程池中执行以避免阻塞）"""
    try:
        file_content = await file.read()
        filename = file.filename

        ext = Path(filename).suffix.lower()
        if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp']:
            resume_text = await asyncio.to_thread(process_image, file_content, filename)
        else:
            resume_text = await asyncio.to_thread(parse_document, file_content, filename)

        if not resume_text or "[解析失败" in resume_text or "[需要" in resume_text:
            return {
                "filename": filename,
                "error": f"简历解析失败: {resume_text}"
            }

        age_info = await asyncio.to_thread(extract_age, resume_text)

        desensitized_text = await asyncio.to_thread(desensitize_resume, resume_text)

        age_min = dimensions_dict.get("age_min") if dimensions_dict else None
        age_max = dimensions_dict.get("age_max") if dimensions_dict else None

        analysis_result = await asyncio.to_thread(
            analyze_resume,
            model_type=current_model,
            api_key=api_key,
            endpoint=endpoint,
            jd_content=jd_content,
            resume_content=desensitized_text,
            resume_name=filename,
            model_name=model_name,
            dimensions=dimensions_dict,
            api_secret=api_secret,
            age_info=age_info,
            age_min=age_min,
            age_max=age_max
        )

        if analysis_result:
            analysis_result = _inject_age_note(analysis_result, age_info, age_min, age_max)
            analysis_result["age_info"] = {
                "age": age_info.get("age"),
                "in_range": _check_age_in_range(age_info, age_min, age_max),
                "source": age_info.get("source")
            }
            return analysis_result
        else:
            return {
                "filename": filename,
                "error": "大模型分析失败"
            }

    except Exception as e:
        return {
            "filename": file.filename if file and file.filename else "unknown",
            "error": str(e)
        }


def _parse_age_val(val: Any) -> Optional[int]:
    try:
        if val in (None, "", "none"):
            return None
        return int(val)
    except (ValueError, TypeError):
        return None


def _check_age_in_range(age_info: Dict[str, Any], age_min: Any, age_max: Any) -> Optional[bool]:
    age = age_info.get("age") if age_info else None
    if age is None:
        return None
    min_i = _parse_age_val(age_min)
    max_i = _parse_age_val(age_max)
    if min_i is not None and age < min_i:
        return False
    if max_i is not None and age > max_i:
        return False
    if min_i is not None or max_i is not None:
        return True
    return None


def _inject_age_note(result: Dict[str, Any], age_info: Dict[str, Any],
                     age_min: Any, age_max: Any) -> Dict[str, Any]:
    """在LLM返回结果后，注入年龄状态备注到shortcomings/matching_points，保证一致性"""
    min_i = _parse_age_val(age_min)
    max_i = _parse_age_val(age_max)

    if min_i is None and max_i is None:
        return result

    age = age_info.get("age") if age_info else None
    shortcomings = result.setdefault("shortcomings", [])
    matching_points = result.setdefault("matching_points", [])

    _AGE_NOTE_PREFIXES = ("⚠️ 年龄", "ℹ️ 简历中未提供年龄", "✅ 年龄")

    shortcomings[:] = [s for s in shortcomings if not any(s.startswith(p) for p in _AGE_NOTE_PREFIXES)]
    matching_points[:] = [m for m in matching_points if not any(m.startswith(p) for p in _AGE_NOTE_PREFIXES)]

    if age is not None:
        in_range = True
        if min_i is not None and age < min_i:
            in_range = False
        if max_i is not None and age > max_i:
            in_range = False

        range_text = ""
        if min_i is not None and max_i is not None:
            range_text = f"{min_i}-{max_i}岁"
        elif min_i is not None:
            range_text = f"{min_i}岁以上"
        elif max_i is not None:
            range_text = f"{max_i}岁以下"

        if in_range:
            matching_points.insert(0, f"✅ 年龄符合要求（{age}岁，在{range_text}范围内）")
        else:
            shortcomings.insert(0, f"⚠️ 年龄不符合要求（要求{range_text}，实际{age}岁）")
    else:
        range_text = ""
        if min_i is not None and max_i is not None:
            range_text = f"{min_i}-{max_i}岁"
        elif min_i is not None:
            range_text = f"{min_i}岁以上"
        elif max_i is not None:
            range_text = f"{max_i}岁以下"
        shortcomings.insert(0, f"ℹ️ 简历中未提供年龄信息，无法判断是否符合{range_text}的年龄段要求")

    return result


def _format_interview_question(index: int, q: Any) -> str:
    """格式化单个面试问题用于 Excel 导出，兼容新旧格式"""
    if isinstance(q, dict):
        question = q.get("question", "")
        focus = q.get("focus", "")
        if focus:
            return f"{index+1}. {question}\n   🎯 考察：{focus}"
        else:
            return f"{index+1}. {question}"
    else:
        return f"{index+1}. {q}"


@app.post("/api/screening")
async def api_screening(
    jd_content: str = Form(...),
    files: List[UploadFile] = File(...),
    dimensions: Optional[str] = Form(None),
    api_key: Optional[str] = Form(None),
    model_type: Optional[str] = Form(None),
    model_endpoint: Optional[str] = Form(None),
    model_name: Optional[str] = Form(None),
    api_secret: Optional[str] = Form(None)
):
    """
    智能筛选主流程接口

    流程：
    1. 解析上传的简历文件
    2. 对图片简历进行OCR识别
    3. 对简历内容进行脱敏
    4. 调用大模型分析匹配度
    5. 返回结构化结果

    dimensions（可选JSON字符串）：
    {
        "work_years": "3年以上",
        "education": "本科及以上",
        "skill_weights": "Python、机器学习优先",
        "extra": "有大厂经验优先"
    }

    生产模式（REQUIRE_USER_API_KEY=true）下，前端必须通过 Form 字段提交
    api_key / model_type / model_endpoint / model_name / api_secret，
    后端不再读取服务器端配置。

    本地模式（默认）下，上述 Form 字段被忽略，沿用 load_config() 行为。
    """
    if REQUIRE_USER_API_KEY:
        if not api_key or not model_endpoint:
            raise HTTPException(
                status_code=400,
                detail="请先在浏览器端完成 API Key 配置"
            )
        current_model = model_type or "deepseek"
        endpoint = model_endpoint
        resolved_model_name = model_name or ""
        resolved_api_secret = api_secret or ""
    else:
        config = load_config()
        if not config.get("is_config_valid", False):
            raise HTTPException(status_code=400, detail="请先完成APIkey配置")
        current_model = config.get("current_model", "deepseek")
        model_config = config.get("models", {}).get(current_model, {})
        api_key = model_config.get("api_key", "")
        endpoint = model_config.get("endpoint", "")
        resolved_model_name = model_config.get("model_version", "")
        resolved_api_secret = model_config.get("api_secret", "")
        if not api_key or not endpoint:
            raise HTTPException(status_code=400, detail="当前模型配置不完整")

    # 解析自定义匹配维度
    dimensions_dict = None
    if dimensions:
        try:
            dimensions_dict = json.loads(dimensions)
        except json.JSONDecodeError:
            dimensions_dict = None

    # 并发处理所有简历（使用 Semaphore 限制并发数）
    semaphore = asyncio.Semaphore(_MAX_CONCURRENT_RESUMES)

    async def _concurrent_process(file: UploadFile):
        async with semaphore:
            return await _process_single_resume(
                file=file,
                current_model=current_model,
                api_key=api_key,
                endpoint=endpoint,
                jd_content=jd_content,
                model_name=resolved_model_name,
                dimensions_dict=dimensions_dict,
                api_secret=resolved_api_secret
            )

    all_results = await asyncio.gather(*[_concurrent_process(f) for f in files])

    # 分离成功结果和错误
    results = []
    errors = []
    for item in all_results:
        if "error" in item:
            errors.append(item)
        else:
            results.append(item)

    # 按匹配度排序（兼容大模型返回字符串型评分）
    def _score_key(x):
        try:
            return int(x.get("score", 0))
        except (TypeError, ValueError):
            return 0
    results.sort(key=_score_key, reverse=True)

    return {
        "success": True,
        "total": len(files),
        "results": results,
        "errors": errors,
        "model_used": current_model,
        "model_name": resolved_model_name or current_model
    }


@app.post("/api/parse-resume")
async def api_parse_resume(file: UploadFile = File(...)):
    """
    单独解析简历接口（用于预览）
    """
    try:
        file_content = await file.read()
        filename = file.filename

        ext = Path(filename).suffix.lower()
        if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp']:
            resume_text = await asyncio.to_thread(process_image, file_content, filename)
        else:
            resume_text = await asyncio.to_thread(parse_document, file_content, filename)

        desensitized_text = await asyncio.to_thread(desensitize_resume, resume_text)

        return {
            "success": True,
            "original_text": resume_text,
            "desensitized_text": desensitized_text,
            "filename": filename
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"简历解析失败: {str(e)}")


# ==================== Excel导出接口 ====================

@app.post("/api/export")
async def api_export_excel(payload: Dict[str, Any]):
    """
    导出筛选结果为Excel文件

    请求体：
    {
        "results": [...],   // 筛选结果数组
        "jd_summary": "xx"  // 可选，JD摘要用作Sheet标题备注
    }
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="导出功能需要安装openpyxl库，请运行: pip install openpyxl")

    results = payload.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="没有可导出的数据")

    try:
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "简历筛选结果"

        # 样式定义
        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cell_align = Alignment(vertical="top", wrap_text=True)
        alt_fill = PatternFill(start_color="F0F0FF", end_color="F0F0FF", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD")
        )

        # 表头
        headers = [
            ("排名", 6),
            ("简历文件名", 25),
            ("匹配度评分", 12),
            ("年龄", 12),
            ("简历概述", 40),
            ("匹配点", 45),
            ("不足点", 35),
            ("建议面试问题", 65),
        ]

        for col_idx, (header_text, col_width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 28

        # 填充数据行
        for row_idx, item in enumerate(results, start=2):
            rank = row_idx - 1
            resume_name = item.get("resume_name", "")
            # 兼容大模型返回字符串型评分的情况
            raw_score = item.get("score", 0)
            try:
                score = int(raw_score)
            except (TypeError, ValueError):
                score = 0
            summary = item.get("summary", "")
            matching_points = "\n".join(
                f"• {p}" for p in (item.get("matching_points") or [])
            )
            shortcomings = "\n".join(
                f"• {s}" for s in (item.get("shortcomings") or [])
            )
            interview_qs = "\n".join(
                _format_interview_question(i, q) for i, q in enumerate(item.get("interview_questions") or [])
            )

            age_info = item.get("age_info") or {}
            age_val = age_info.get("age")
            in_range = age_info.get("in_range")
            if age_val is None:
                age_display = "未知"
            elif in_range is True:
                age_display = f"{age_val}岁 ✓"
            elif in_range is False:
                age_display = f"{age_val}岁 ✗"
            else:
                age_display = f"{age_val}岁"

            row_data = [rank, resume_name, score, age_display, summary, matching_points, shortcomings, interview_qs]
            use_alt = (row_idx % 2 == 0)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
                cell.border = thin_border
                if use_alt:
                    cell.fill = alt_fill
                if col_idx in (3, 4):
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    cell.font = Font(bold=True, size=12)

            # 自动行高（近似估算）
            max_lines = max(
                len(str(v).split("\n")) for v in row_data if v
            ) if any(row_data) else 1
            ws.row_dimensions[row_idx].height = max(18, min(max_lines * 16, 200))

        # 冻结首行
        ws.freeze_panes = "A2"

        # 输出到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import datetime
        from urllib.parse import quote
        filename = f"简历筛选结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # RFC 5987 要求 filename*=UTF-8'' 后跟 percent-encoded 文件名，
        # 不能直接放中文（HTTP header 仅允许 ASCII，否则会触发 UnicodeEncodeError）
        encoded_filename = quote(filename)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


# ==================== 健康检查 ====================

@app.get("/api/health")
async def health_check():
    """健康检查"""
    return {
        "status": "ok",
        "version": "1.2.0"
    }


# ==================== 启动配置 ====================

if __name__ == "__main__":
    # 确保配置目录存在
    config_dir = ROOT_DIR / "config"
    config_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 50)
    print("智能简历筛选工具 V1.2")
    print("=" * 50)
    print(f"前端目录: {FRONTEND_DIR}")
    print(f"配置目录: {config_dir}")
    print("=" * 50)
    print("服务启动成功！")
    print("访问地址: http://localhost:8000")
    print("=" * 50)

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
